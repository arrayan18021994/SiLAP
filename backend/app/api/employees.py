from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import io
import datetime
from openpyxl import load_workbook
from fastapi.responses import StreamingResponse

from app.database.session import get_db
from app.database.models import Employee, Organization, Position, Rank
from app.services.employee_import_service import generate_employee_template
from app.services.family_service import FamilyService
from app.api.auth import get_current_user

router = APIRouter()

@router.get("/template")
def download_template():
    """Download standard employee import Excel template with headers and example rows."""
    file_bytes = generate_employee_template()
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=SiLAP_Template_Pegawai.xlsx"}
    )

def get_rank_score(rank_str: str) -> int:
    if not rank_str or rank_str == '-':
        return 0
    r = str(rank_str).strip().upper()
    scores = {
        "IV/E": 45, "IV/D": 44, "IV/C": 43, "IV/B": 42, "IV/A": 41,
        "III/D": 34, "III/C": 33, "III/B": 32, "III/A": 31,
        "II/D": 24, "II/C": 23, "II/B": 22, "II/A": 21,
        "I/D": 14, "I/C": 13, "I/B": 12, "I/A": 11,
    }
    if r in scores:
        return scores[r]
    roman_scores = {
        "XVII": 17, "XVI": 16, "XV": 15, "XIV": 14, "XIII": 13, "XII": 12, "XI": 11,
        "X": 10, "IX": 9, "VIII": 8, "VII": 7, "VI": 6, "V": 5, "IV": 4, "III": 3, "II": 2, "I": 1
    }
    for k, v in roman_scores.items():
        if k in r:
            return v
    return 0

def get_asn_status_score(status_str: str) -> int:
    if not status_str:
        return 99
    st = str(status_str).strip().upper()
    if "PARUH" in st or "PART" in st:
        return 3
    if "PENUH" in st or "FULL" in st or "PPPK" in st:
        return 2
    if "PNS" in st:
        return 1
    return 4

@router.get("/")
def get_employees(skip: int = 0, limit: int = 200, db: Session = Depends(get_db)):
    """Retrieve employees sorted hierarchically by ASN Status, Golongan, MKG, and NIP."""
    from app.database.models.family import FamilyMember
    employees = db.query(Employee).filter(Employee.is_active == True).all()
    result = []
    for emp in employees:
        children_cnt = db.query(FamilyMember).filter(
            FamilyMember.employee_id == emp.id,
            FamilyMember.relationship_type.ilike("%ANAK%")
        ).count()

        result.append({
            "id": emp.id,
            "nip": emp.nip,
            "name": getattr(emp, 'full_name', getattr(emp, 'name', '')),
            "full_name": getattr(emp, 'full_name', getattr(emp, 'name', '')),
            "nik": emp.nik,
            "birth_place": emp.birth_place,
            "birth_date": str(emp.birth_date) if emp.birth_date else None,
            "gender": emp.gender,
            "status": emp.asn_status if emp.asn_status else "PNS",
            "asn_status": emp.asn_status if emp.asn_status else "PNS",
            "marital_status": emp.marital_status or "BELUM KAWIN",
            "mkg_years": getattr(emp, 'mkg_years', 0) or 0,
            "mkg_months": getattr(emp, 'mkg_months', 0) or 0,
            "tmt_mkg": str(emp.tmt_mkg) if getattr(emp, 'tmt_mkg', None) else None,
            "children_count": children_cnt,
            "address": emp.address,
            "notes": emp.notes,
            "rank": emp.rank if getattr(emp, 'rank', None) else '-',
            "position": emp.position if getattr(emp, 'position', None) else '-',
            "opd": emp.opd if getattr(emp, 'opd', None) else '-'
        })

    # Sort hierarchically: Status ASN (1,2,3) -> Rank (45..0 DESC) -> MKG Months (DESC) -> NIP (ASC)
    result.sort(key=lambda x: (
        get_asn_status_score(x.get("status") or x.get("asn_status")),
        -get_rank_score(x.get("rank")),
        -((int(x.get("mkg_years") or 0) * 12) + int(x.get("mkg_months") or 0)),
        str(x.get("nip") or "")
    ))
    return result[skip:skip+limit]

@router.get("/{employee_id}")
def get_employee(employee_id: int, db: Session = Depends(get_db)):
    from app.database.models.family import FamilyMember
    emp = db.query(Employee).filter(Employee.id == employee_id, Employee.is_active == True).first()
    if not emp:
         raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")

    children_cnt = db.query(FamilyMember).filter(
        FamilyMember.employee_id == emp.id,
        FamilyMember.relationship_type.ilike("%ANAK%")
    ).count()

    return {
        "id": emp.id,
        "nip": emp.nip,
        "name": getattr(emp, 'full_name', getattr(emp, 'name', '')),
        "full_name": getattr(emp, 'full_name', getattr(emp, 'name', '')),
        "nik": emp.nik,
        "birth_place": emp.birth_place,
        "birth_date": str(emp.birth_date) if emp.birth_date else None,
        "gender": emp.gender,
        "status": emp.asn_status,
        "asn_status": emp.asn_status,
        "marital_status": emp.marital_status or "BELUM KAWIN",
        "mkg_years": getattr(emp, 'mkg_years', 0) or 0,
        "mkg_months": getattr(emp, 'mkg_months', 0) or 0,
        "tmt_mkg": str(emp.tmt_mkg) if getattr(emp, 'tmt_mkg', None) else None,
        "children_count": children_cnt,
        "address": emp.address,
        "notes": emp.notes,
        "rank": getattr(emp, 'rank', '-'),
        "position": getattr(emp, 'position', '-'),
        "opd": getattr(emp, 'opd', '-')
    }

@router.post("/")
def create_employee(data: dict, db: Session = Depends(get_db)):
    nip = data.get("nip")
    if not nip:
        raise HTTPException(status_code=400, detail="NIP wajib diisi.")
        
    if db.query(Employee).filter(Employee.nip == nip).first():
        raise HTTPException(status_code=400, detail="NIP sudah terdaftar.")
    
    full_name = data.get("full_name") or data.get("name") or ""
    asn_status = data.get("asn_status") or data.get("status") or "PNS"
    
def parse_date_str(d_str):
    if not d_str:
        return None
    d_str_str = str(d_str).strip()
    if isinstance(d_str, (datetime.date, datetime.datetime)):
        return d_str if isinstance(d_str, datetime.date) else d_str.date()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(d_str_str, fmt).date()
        except ValueError:
            pass
    return None

def format_date_str(d):
    if not d:
        return None
    if isinstance(d, (datetime.date, datetime.datetime)):
        return d.strftime("%d/%m/%Y")
    parsed = parse_date_str(d)
    return parsed.strftime("%d/%m/%Y") if parsed else str(d)



@router.post("/")
def create_employee(data: dict, db: Session = Depends(get_db)):
    nip = data.get("nip")
    if not nip:
        raise HTTPException(status_code=400, detail="NIP wajib diisi.")
        
    if db.query(Employee).filter(Employee.nip == nip).first():
        raise HTTPException(status_code=400, detail="NIP sudah terdaftar.")
    
    full_name = data.get("full_name") or data.get("name") or ""
    asn_status = data.get("asn_status") or data.get("status") or "PNS"
    
    birth_date = parse_date_str(data.get("birth_date"))
    tmt_cpns = parse_date_str(data.get("tmt_cpns"))
    tmt_mkg = parse_date_str(data.get("tmt_mkg"))

    try:
        mkg_years = int(data.get("mkg_years", 0))
    except (ValueError, TypeError):
        mkg_years = 0

    try:
        mkg_months = int(data.get("mkg_months", 0))
    except (ValueError, TypeError):
        mkg_months = 0

    emp = Employee(
        nip=nip,
        full_name=full_name,
        nik=data.get("nik"),
        birth_place=data.get("birth_place"),
        birth_date=birth_date,
        gender=data.get("gender"),
        asn_status=asn_status,
        employment_status=data.get("employment_status", "Aktif"),
        marital_status=data.get("marital_status"),
        phone=data.get("phone"),
        email=data.get("email"),
        address=data.get("address"),
        tmt_cpns=tmt_cpns,
        mkg_years=mkg_years,
        mkg_months=mkg_months,
        tmt_mkg=tmt_mkg,
        notes=data.get("notes")
    )
    
    # Store extra optional fields if present as model attributes
    for extra_field in ["name", "status", "rank", "position", "opd", "unit_kerja"]:
        if extra_field in data:
            setattr(emp, extra_field, data[extra_field])
            
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp

@router.put("/{employee_id}")
def update_employee(employee_id: int, data: dict, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == employee_id, Employee.is_active == True).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
        
    nip = data.get("nip")
    if nip and nip != emp.nip:
        existing = db.query(Employee).filter(Employee.nip == nip, Employee.id != employee_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="NIP sudah digunakan oleh pegawai lain.")
        emp.nip = nip

    if "full_name" in data or "name" in data:
        emp.full_name = data.get("full_name") or data.get("name") or emp.full_name
    if "nik" in data:
        emp.nik = data["nik"]
    if "birth_place" in data:
        emp.birth_place = data["birth_place"]
    if "birth_date" in data:
        emp.birth_date = parse_date_str(data["birth_date"])
    if "gender" in data:
        emp.gender = data["gender"]
    if "asn_status" in data or "status" in data:
        emp.asn_status = data.get("asn_status") or data.get("status") or emp.asn_status
    if "marital_status" in data:
        emp.marital_status = data["marital_status"]
    if "address" in data:
        emp.address = data["address"]
    if "notes" in data:
        emp.notes = data["notes"]
    if "tmt_cpns" in data:
        emp.tmt_cpns = parse_date_str(data["tmt_cpns"])
    if "tmt_mkg" in data:
        emp.tmt_mkg = parse_date_str(data["tmt_mkg"])
    if "mkg_years" in data:
        try:
            emp.mkg_years = int(data["mkg_years"])
        except (ValueError, TypeError):
            pass
    if "mkg_months" in data:
        try:
            emp.mkg_months = int(data["mkg_months"])
        except (ValueError, TypeError):
            pass

    for extra_field in ["name", "status", "rank", "position", "opd", "unit_kerja"]:
        if extra_field in data:
            setattr(emp, extra_field, data[extra_field])

    db.commit()
    db.refresh(emp)
    return emp

@router.patch("/{employee_id}/status")
def deactivate_employee(employee_id: int, db: Session = Depends(get_db)):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
         raise HTTPException(status_code=404, detail="Pegawai tidak ditemukan")
         
    emp.is_active = False
    db.commit()
    return {"message": "Pegawai dinonaktifkan."}

@router.post("/preview")
async def preview_employees_import(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File harus berformat .xlsx")
        
    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    
    sheet_name = None
    for target in ["02_DATA_PEGAWAI", "Data_Pegawai"]:
        if target in wb.sheetnames:
            sheet_name = target
            break
            
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
        
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    
    if len(rows) < 2:
        return {"error": "Data kosong."}
        
    data_rows = rows[1:]
    
    children_by_nip = {}
    if "04_DATA_ANAK" in wb.sheetnames:
        child_sheet = wb["04_DATA_ANAK"]
        child_rows = list(child_sheet.iter_rows(values_only=True))
        if len(child_rows) > 1:
            for crow in child_rows[1:]:
                if not crow or crow[0] is None:
                    continue
                cnip = str(crow[0]).strip()
                if cnip.upper().startswith("NIP") or cnip.upper().startswith("CONTOH"):
                    continue
                if cnip not in children_by_nip:
                    children_by_nip[cnip] = []
                children_by_nip[cnip].append({
                    "name": str(crow[1]).strip() if len(crow) > 1 and crow[1] is not None else "",
                    "nik": str(crow[2]).strip() if len(crow) > 2 and crow[2] is not None else None,
                    "birth_place": str(crow[3]).strip() if len(crow) > 3 and crow[3] is not None else None,
                    "birth_date": str(crow[4]).strip() if len(crow) > 4 and crow[4] is not None else None,
                    "gender": str(crow[5]).strip() if len(crow) > 5 and crow[5] is not None else None,
                    "document_number": str(crow[6]).strip() if len(crow) > 6 and crow[6] is not None else None,
                    "child_status": str(crow[7]).strip() if len(crow) > 7 and crow[7] is not None else "Anak Kandung",
                    "education": str(crow[8]).strip() if len(crow) > 8 and crow[8] is not None else None,
                })

    preview_data = []
    
    for idx, row in enumerate(data_rows, start=2):
        if not row or all(cell is None for cell in row):
            continue
            
        nip_val = str(row[0]).strip() if row[0] is not None else ""
        if nip_val.upper().startswith("NIP") or nip_val.upper().startswith("CONTOH"):
            continue
            
        name_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        
        status_msg = "VALID"
        messages = []
        if not nip_val:
            status_msg = "ERROR"
            messages.append("NIP kosong")
        if not name_val:
            status_msg = "ERROR"
            messages.append("Nama kosong")
            
        existing = db.query(Employee).filter(Employee.nip == nip_val).first() if nip_val else None
        if existing:
            status_msg = "ERROR"
            messages.append("NIP sudah terdaftar")
            
        # Parse fields dynamically based on row length and headers
        nik_val = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        birth_place_val = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        birth_date_val = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        gender_val = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        asn_status_val = str(row[6]).strip() if len(row) > 6 and row[6] else "PNS"
        tmt_asn_val = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        rank_val = str(row[8]).strip() if len(row) > 8 and row[8] else "-"
        
        # Check if 19-column layout (with MKG columns 10, 11, 12)
        mkg_years_val = 0
        mkg_months_val = 0
        tmt_mkg_val = ""
        position_val = "-"
        opd_val = "-"
        marital_val = ""
        address_val = ""
        notes_val = ""

        if len(row) >= 18 or (len(row) > 10 and str(row[10]).isdigit()):
            try:
                mkg_years_val = int(row[10]) if row[10] is not None and str(row[10]).isdigit() else 0
            except Exception:
                mkg_years_val = 0
            try:
                mkg_months_val = int(row[11]) if len(row) > 11 and row[11] is not None and str(row[11]).isdigit() else 0
            except Exception:
                mkg_months_val = 0
            tmt_mkg_val = str(row[12]).strip() if len(row) > 12 and row[12] else ""
            position_val = str(row[13]).strip() if len(row) > 13 and row[13] else "-"
            opd_val = str(row[15]).strip() if len(row) > 15 and row[15] else "-"
            address_val = str(row[16]).strip() if len(row) > 16 and row[16] else ""
            marital_val = str(row[17]).strip() if len(row) > 17 and row[17] else ""
            notes_val = str(row[18]).strip() if len(row) > 18 and row[18] else ""
        else:
            position_val = str(row[10]).strip() if len(row) > 10 and row[10] else "-"
            opd_val = str(row[12]).strip() if len(row) > 12 and row[12] else "-"
            address_val = str(row[14]).strip() if len(row) > 14 and row[14] else ""
            marital_val = str(row[15]).strip() if len(row) > 15 and row[15] else ""
            notes_val = str(row[16]).strip() if len(row) > 16 and row[16] else ""

        preview_data.append({
            "row": idx,
            "nip": nip_val,
            "name": name_val,
            "full_name": name_val,
            "nik": nik_val,
            "birth_place": birth_place_val,
            "birth_date": birth_date_val,
            "gender": gender_val,
            "status": asn_status_val,
            "asn_status": asn_status_val,
            "tmt_asn": tmt_asn_val,
            "tmt_cpns": tmt_asn_val,
            "rank": rank_val,
            "mkg_years": mkg_years_val,
            "mkg_months": mkg_months_val,
            "tmt_mkg": tmt_mkg_val,
            "position": position_val,
            "opd": opd_val,
            "marital_status": marital_val,
            "address": address_val,
            "notes": notes_val,
            "children": children_by_nip.get(nip_val, []),
            "import_status": status_msg,
            "messages": messages
        })
        
    return {
        "total_rows": len(preview_data),
        "valid_rows": sum(1 for p in preview_data if p["import_status"] == "VALID"),
        "error_rows": sum(1 for p in preview_data if p["import_status"] == "ERROR"),
        "preview_data": preview_data
    }

@router.post("/commit")
async def commit_employees_import(data: dict, db: Session = Depends(get_db)):
    preview_data = data.get("preview_data", [])
    imported = 0
    for row in preview_data:
        if row.get("import_status") == "VALID":
            birth_date = parse_date_str(row.get("birth_date"))
            tmt_cpns = parse_date_str(row.get("tmt_cpns") or row.get("tmt_asn"))
            tmt_mkg = parse_date_str(row.get("tmt_mkg"))
            
            emp = Employee(
                nip=row["nip"],
                full_name=row.get("name") or row.get("full_name") or "",
                nik=row.get("nik"),
                birth_place=row.get("birth_place"),
                birth_date=birth_date,
                gender=row.get("gender"),
                asn_status=row.get("asn_status") or row.get("status") or "PNS",
                employment_status="Aktif",
                marital_status=row.get("marital_status"),
                address=row.get("address"),
                tmt_cpns=tmt_cpns,
                mkg_years=row.get("mkg_years", 0),
                mkg_months=row.get("mkg_months", 0),
                tmt_mkg=tmt_mkg,
                notes=row.get("notes")
            )
            for extra in ["name", "status", "rank", "position", "opd"]:
                if extra in row:
                    setattr(emp, extra, row[extra])
            db.add(emp)
            db.flush()
            
            for child in row.get("children", []):
                if child.get("name"):
                    FamilyService.add_family_member(db, emp.id, {
                        "relationship_type": "ANAK",
                        "name": child["name"],
                        "nik": child.get("nik"),
                        "birth_place": child.get("birth_place"),
                        "birth_date": child.get("birth_date"),
                        "gender": child.get("gender"),
                        "document_number": child.get("document_number"),
                        "child_status": child.get("child_status"),
                        "education": child.get("education")
                    })

            imported += 1
    
    db.commit()
    return {"message": "Import berhasil", "imported_rows": imported}


