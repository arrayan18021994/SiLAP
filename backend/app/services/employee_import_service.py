import io
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill

def generate_employee_template():
    wb = Workbook()
    
    # 01_PETUNJUK
    ws1 = wb.active
    ws1.title = "01_PETUNJUK"
    ws1.append(["PETUNJUK PENGISIAN TEMPLATE DATA PEGAWAI"])
    ws1.append(["1. Jangan mengubah nama sheet atau urutan kolom."])
    ws1.append(["2. Kolom bertanda bintang (*) adalah kolom wajib diisi."])
    ws1.append(["3. NIP adalah identitas unik pegawai (18 digit angka)."])
    ws1.append(["4. Hapus baris contoh sebelum mengunggah file untuk import data sesungguhnya."])
    ws1.append(["5. Gunakan format tanggal standar DD/MM/YYYY (contoh: 10/05/1985 atau 01/04/2008)."])
    
    ws1.column_dimensions['A'].width = 80
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="1E3A8A")

    # 02_DATA_PEGAWAI
    ws2 = wb.create_sheet(title="02_DATA_PEGAWAI")
    headers2 = [
        "NIP *", "Nama & Gelar *", "NIK", "Tempat Lahir *", "Tanggal Lahir *", 
        "Jenis Kelamin *", "Status ASN *", "TMT ASN *", "Pangkat/Golongan", 
        "TMT Pangkat", "MKG (Tahun)", "MKG (Bulan)", "TMT MKG", "Jabatan", "TMT Jabatan", "Unit Kerja", 
        "Alamat", "Status Perkawinan *"
    ]
    ws2.append(headers2)
    
    # Sample Rows
    ws2.append([
        "198505102008041001", "Ahmad Budi, S.E.", "1172011005850001", "Sabang", "10/05/1985",
        "L", "PNS", "01/04/2008", "III/a", "01/04/2008",
        5, 2, "01/04/2018",
        "Analis Kepegawaian", "01/04/2008", "Bidang Mutasi", "Jl. Yos Sudarso No. 12", "KAWIN"
    ])
    ws2.append([
        "199008152022212003", "Siti Rahmah, S.Pd.", "1172025508900002", "Banda Aceh", "15/08/1990",
        "P", "PPPK Penuh Waktu", "01/02/2022", "IX", "01/02/2022",
        2, 0, "01/02/2022",
        "Guru Ahli Pertama", "01/02/2022", "SMP Negeri 1", "Jl. T. Umar No. 45", "KAWIN"
    ])
    
    # 03_DATA_PASANGAN
    ws3 = wb.create_sheet(title="03_DATA_PASANGAN")
    headers3 = ["NIP Pegawai *", "Nama Pasangan", "NIK Pasangan", "Tempat Lahir", "Tanggal Lahir", "Tanggal Perkawinan", "Pekerjaan"]
    ws3.append(headers3)
    ws3.append(["198505102008041001", "Siti Aminah", "1172015206880003", "Banda Aceh", "12/06/1988", "15/08/2010", "Wiraswasta"])
    
    # 04_DATA_ANAK
    ws4 = wb.create_sheet(title="04_DATA_ANAK")
    headers4 = ["NIP Pegawai *", "Nama Anak", "NIK Anak", "Tempat Lahir", "Tanggal Lahir", "Jenis Kelamin", "Nomor Akta Lahir", "Status Anak", "Pendidikan"]
    ws4.append(headers4)
    ws4.append(["198505102008041001", "Budi Ahmad", "1172010101150005", "Sabang", "01/01/2015", "L", "1172-LU-00001", "Anak Kandung", "SD/Sederajat"])
    
    # 05_REFERENSI
    ws5 = wb.create_sheet(title="05_REFERENSI")
    ws5.append(["STATUS ASN", "JENIS KELAMIN", "STATUS PERKAWINAN"])
    ws5.append(["PNS", "L", "BELUM KAWIN"])
    ws5.append(["PPPK Penuh Waktu", "P", "KAWIN"])
    ws5.append(["PPPK Paruh Waktu", "", "CERAI HIDUP"])
    ws5.append(["", "", "CERAI MATI"])
    
    # Data Validation
    dv_status = DataValidation(type="list", formula1="='05_REFERENSI'!$A$2:$A$4", allow_blank=True)
    dv_gender = DataValidation(type="list", formula1="='05_REFERENSI'!$B$2:$B$3", allow_blank=True)
    dv_marital = DataValidation(type="list", formula1="='05_REFERENSI'!$C$2:$C$5", allow_blank=True)
    
    ws2.add_data_validation(dv_status)
    ws2.add_data_validation(dv_gender)
    ws2.add_data_validation(dv_marital)
    
    dv_status.add("G2:G1000") # Status ASN
    dv_gender.add("F2:F1000") # Jenis Kelamin
    dv_marital.add("R2:R1000") # Status Perkawinan (Col 18)
    
    # Styling headers & columns
    header_fill = PatternFill(start_color="FACC15", end_color="FACC15", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for ws in [ws2, ws3, ws4, ws5]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
