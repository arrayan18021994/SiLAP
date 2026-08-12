import io
import uuid
import datetime
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database.models import Employee, EmployeeServicePeriod, ImportBatch, ImportError
from app.services.service_period import ServicePeriodCalculationService

class EmployeeImportService:

    @staticmethod
    def preview_service_period_import(db: Session, file_content: bytes, user_id: str, file_name: str) -> dict:
        """Parses the Excel file using openpyxl, validates rows, and returns a preview without saving to DB."""
        wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
        if "DATA_MASA_KERJA" not in wb.sheetnames:
            return {"error": "Sheet 'DATA_MASA_KERJA' tidak ditemukan dalam file Excel."}
            
        sheet = wb["DATA_MASA_KERJA"]
        rows = list(sheet.iter_rows(values_only=True))
        
        if len(rows) < 2:
            return {"error": "Data kosong."}
            
        header = rows[0]
        data_rows = rows[1:]
        
        batch_id = str(uuid.uuid4())
        valid_rows = 0
        error_rows = 0
        errors = []
        preview_data = []

        for idx, row in enumerate(data_rows, start=2): # 1-based index, row 1 is header
            if not row or all(cell is None for cell in row):
                continue
                
            # Column mapping based on template:
            # 0: No, 1: NIP, 2: Nama, 3: TMT CPNS, 4: TMT PNS, 5: Base Year, 6: Base Month, 
            # 7: Adj Year, 8: Adj Month, 9: Eff Date, 10: Doc Num, 11: Doc Date, 12: Notes
            nip = str(row[1]).strip() if row[1] else None
            
            row_errors = []
            if not nip:
                row_errors.append("NIP kosong.")
            else:
                emp = db.query(Employee).filter(Employee.nip == nip).first()
                if not emp:
                    row_errors.append(f"Pegawai dengan NIP {nip} tidak ditemukan.")
            
            try:
                base_y = int(row[5] or 0)
                base_m = int(row[6] or 0)
                adj_y = int(row[7] or 0)
                adj_m = int(row[8] or 0)
            except ValueError:
                row_errors.append("Masa kerja tahun/bulan harus berupa angka.")
                
            eff_date = row[9]
            if not eff_date:
                row_errors.append("Tanggal Efektif kosong.")
            elif not isinstance(eff_date, datetime.datetime) and not isinstance(eff_date, datetime.date):
                 # simplistic check, openpyxl usually returns datetime if formatted as date
                 row_errors.append("Format Tanggal Efektif tidak valid.")

            if row_errors:
                error_rows += 1
                for err in row_errors:
                    errors.append({
                        "row_number": idx,
                        "field_name": "Multiple",
                        "message": err,
                        "severity": "ERROR"
                    })
                preview_data.append({"row": idx, "nip": nip, "status": "ERROR", "messages": row_errors})
            else:
                valid_rows += 1
                preview_data.append({"row": idx, "nip": nip, "status": "VALID", "messages": []})
                
        return {
            "batch_id": batch_id,
            "total_rows": valid_rows + error_rows,
            "valid_rows": valid_rows,
            "error_rows": error_rows,
            "errors": errors,
            "preview_data": preview_data,
            "file_name": file_name
        }

    # Similar method for commit, skipping for brevity in mock
